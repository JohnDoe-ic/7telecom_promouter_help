import logging
import asyncio
from aiogram import Bot, Dispatcher, types, F, Router  # Импортируем Router
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.filters import CommandStart, Command
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder
from aiogram.client.default import DefaultBotProperties  # Для DeprecationWarning
import openpyxl
from datetime import datetime, timedelta
import os

# Замените 'YOUR_BOT_TOKEN' на токен вашего бота
API_TOKEN = '8529869959:AAFuNK1WoYQ0lJ7S8kXLVbIIFlKDp65piCw'
# Замените 'YOUR_MANAGER_ID' на Telegram ID руководителя (пример)
# Важно: MANAGER_ID должен быть числом!
MANAGER_ID = 521620770

# Путь к файлу Excel
EXCEL_FILE = 'promoters_report.xlsx'

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(name)s - %(message)s')
logger = logging.getLogger(__name__)


# FSM для промоутера
class PromoterState(StatesGroup):
    waiting_for_name = State()
    waiting_for_address = State()
    waiting_for_work_time = State()


# FSM для руководителя (для выбора периода отчета/выплат)
class ManagerReportState(StatesGroup):
    waiting_for_report_period = State()


# --- Вспомогательные функции для работы с Excel ---

def init_excel():
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Отчеты промоутеров"
        sheet['A1'] = "Дата"
        sheet['B1'] = "Время"
        sheet['C1'] = "ID промоутера"
        sheet['D1'] = "Имя промоутера"
        sheet['E1'] = "Адрес работы"
        sheet['F1'] = "Планируемое время работы"
        sheet['G1'] = "Статус"  # Ожидает, Подтвержден, Отклонен
        sheet['H1'] = "ID записи"  # Уникальный ID для каждой записи
        workbook.save(EXCEL_FILE)
    return workbook


def add_report_to_excel(promoter_id, promoter_name, address, work_time, record_id):
    workbook = init_excel()
    sheet = workbook["Отчеты промоутеров"]
    now = datetime.now()
    sheet.append([
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        promoter_id,
        promoter_name,
        address,
        work_time,
        "Ожидает",
        record_id
    ])
    workbook.save(EXCEL_FILE)
    logger.info(f"Report added for promoter {promoter_id} (record_id: {record_id})")


def update_report_status(record_id, status):
    workbook = init_excel()
    sheet = workbook["Отчеты промоутеров"]
    found = False
    for row_index in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=8).value == record_id:
            sheet.cell(row=row_index, column=7).value = status
            found = True
            break
    workbook.save(EXCEL_FILE)
    if found:
        logger.info(f"Report {record_id} status updated to {status}")
    else:
        logger.warning(f"Report with record_id {record_id} not found for status update.")


def get_reports_for_period(start_date, end_date):
    workbook = init_excel()
    sheet = workbook["Отчеты промоутеров"]
    reports = []
    for row_index in range(2, sheet.max_row + 1):
        try:
            report_date_str = sheet.cell(row=row_index, column=1).value
            # Проверяем, что report_date_str не None и является строкой
            if report_date_str and isinstance(report_date_str, str):
                report_date = datetime.strptime(report_date_str, "%Y-%m-%d").date()
                if start_date <= report_date <= end_date:
                    reports.append({
                        "Дата": sheet.cell(row=row_index, column=1).value,
                        "Время": sheet.cell(row=row_index, column=2).value,
                        "ID промоутера": sheet.cell(row=row_index, column=3).value,
                        "Имя промоутера": sheet.cell(row=row_index, column=4).value,
                        "Адрес работы": sheet.cell(row=row_index, column=5).value,
                        "Планируемое время работы": sheet.cell(row=row_index, column=6).value,
                        "Статус": sheet.cell(row=row_index, column=7).value,
                        "ID записи": sheet.cell(row=row_index, column=8).value,
                    })
        except (ValueError, TypeError) as e:
            logger.warning(
                f"Skipping row {row_index} due to date parsing error: {e}. Data: {sheet.cell(row=row_index, column=1).value}")
            continue
    return reports


def calculate_payments(promoters_reports):
    payments = {}
    for report in promoters_reports:
        if report["Статус"] == "Подтвержден":
            promoter_name = report["Имя промоутера"]

            if promoter_name not in payments:
                payments[promoter_name] = 0

            payment_per_shift = 1000  # Допустим, оплата за одну смену 1000 руб.
            payments[promoter_name] += payment_per_shift
    return payments


# --- Клавиатуры ---

def get_promoter_keyboard():
    builder = ReplyKeyboardBuilder()
    builder.add(types.KeyboardButton(text="Сообщить о выходе"))
    return builder.as_markup(resize_keyboard=True)


def get_manager_keyboard():
    builder = ReplyKeyboardBuilder()
    builder.add(types.KeyboardButton(text="Отчет"))
    builder.add(types.KeyboardButton(text="Выплаты"))
    return builder.as_markup(resize_keyboard=True)


def get_manager_report_period_keyboard():
    builder = InlineKeyboardBuilder()
    builder.button(text="За сегодня", callback_data="report_today")
    builder.button(text="За неделю", callback_data="report_week")
    builder.button(text="За месяц", callback_data="report_month")
    builder.adjust(2)
    return builder.as_markup()


def get_manager_payments_period_keyboard():
    builder = InlineKeyboardBuilder()
    builder.button(text="За сегодня", callback_data="payments_today")
    builder.button(text="За неделю", callback_data="payments_week")
    builder.button(text="За месяц", callback_data="payments_month")
    builder.adjust(2)
    return builder.as_markup()


# --- Роутер для всех обработчиков ---
router = Router()


@router.message(CommandStart())
async def cmd_start(message: types.Message, state: FSMContext):
    await state.clear()
    if message.from_user.id == MANAGER_ID:
        await message.reply("Привет, руководитель! Выберите действие:", reply_markup=get_manager_keyboard())
    else:
        await message.reply("Привет! Я бот для учета твоих выходов на смену.", reply_markup=get_promoter_keyboard())


@router.message(F.text == "Сообщить о выходе")
@router.message(Command("report"))
async def start_report_flow(message: types.Message, state: FSMContext):
    await state.set_state(PromoterState.waiting_for_name)
    await message.reply("Отлично! Напиши свое ФИО.", reply_markup=types.ReplyKeyboardRemove())


@router.message(PromoterState.waiting_for_name)
async def process_name(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text)
    await state.set_state(PromoterState.waiting_for_address)
    await message.reply("Теперь напиши адрес, где будешь работать.")


@router.message(PromoterState.waiting_for_address)
async def process_address(message: types.Message, state: FSMContext):
    await state.update_data(address=message.text)
    await state.set_state(PromoterState.waiting_for_work_time)
    await message.reply("Напиши планируемое время работы (например, 'с 10:00 до 18:00').")


@router.message(PromoterState.waiting_for_work_time)
async def process_work_time(message: types.Message, state: FSMContext, bot: Bot):
    user_data = await state.get_data()
    promoter_name = user_data['name']
    promoter_address = user_data['address']
    promoter_work_time = message.text
    promoter_id = message.from_user.id

    record_id = f"{promoter_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"

    add_report_to_excel(promoter_id, promoter_name, promoter_address, promoter_work_time, record_id)

    await message.reply("Спасибо! Твой отчет отправлен руководителю на подтверждение.",
                        reply_markup=get_promoter_keyboard())

    confirmation_builder = InlineKeyboardBuilder()
    confirmation_builder.button(text="Подтвердить", callback_data=f"confirm_{record_id}")
    confirmation_builder.button(text="Отклонить", callback_data=f"decline_{record_id}")
    confirmation_builder.adjust(2)

    await bot.send_message(
        MANAGER_ID,
        f"Новый отчет от промоутера:\n"
        f"ФИО: {promoter_name}\n"
        f"Адрес: {promoter_address}\n"
        f"Время работы: {promoter_work_time}\n"
        f"ID записи: {record_id}",
        reply_markup=confirmation_builder.as_markup()
    )

    await state.clear()


# --- Обработчики для руководителя ---

@router.message(F.text == "Отчет", F.from_user.id == MANAGER_ID)
async def manager_get_report_prompt(message: types.Message, state: FSMContext):
    await state.set_state(ManagerReportState.waiting_for_report_period)
    await message.reply("За какой период вы хотите получить отчет?",
                        reply_markup=get_manager_report_period_keyboard())


@router.message(F.text == "Выплаты", F.from_user.id == MANAGER_ID)
async def manager_calculate_payments_prompt(message: types.Message, state: FSMContext):
    await state.set_state(ManagerReportState.waiting_for_report_period)
    await message.reply("За какой период вы хотите рассчитать выплаты?",
                        reply_markup=get_manager_payments_period_keyboard())


@router.callback_query(lambda c: c.data.startswith('confirm_') or c.data.startswith('decline_'))
async def process_manager_decision(callback_query: types.CallbackQuery, bot: Bot):
    action, record_id = callback_query.data.split('_', 1)

    if action == "confirm":
        update_report_status(record_id, "Подтвержден")
        await bot.edit_message_text(
            chat_id=callback_query.message.chat.id,
            message_id=callback_query.message.message_id,
            text=callback_query.message.text + "\n\n✅ Статус: Подтвержден"
        )
        await callback_query.answer("Отчет подтвержден!")
    elif action == "decline":
        update_report_status(record_id, "Отклонен")
        await bot.edit_message_text(
            chat_id=callback_query.message.chat.id,
            message_id=callback_query.message.message_id,
            text=callback_query.message.text + "\n\n❌ Статус: Отклонен"
        )
        await callback_query.answer("Отчет отклонен!")


@router.callback_query(ManagerReportState.waiting_for_report_period, F.data.startswith(('report_', 'payments_')),
                       F.from_user.id == MANAGER_ID)
async def manager_send_report_or_payments(callback_query: types.CallbackQuery, state: FSMContext, bot: Bot):
    command_type, period = callback_query.data.split('_')

    end_date = datetime.now().date()
    if period == "today":
        start_date = end_date
    elif period == "week":
        start_date = end_date - timedelta(days=6)
    elif period == "month":
        start_date = end_date - timedelta(days=29)

    if command_type == "report":
        reports = get_reports_for_period(start_date, end_date)

        if not reports:
            await callback_query.message.answer(f"За выбранный период ({start_date} - {end_date}) отчетов не найдено.",
                                                reply_markup=get_manager_keyboard())
        else:
            report_workbook = openpyxl.Workbook()
            report_sheet = report_workbook.active
            report_sheet.title = "Отчет"
            report_sheet.append(
                ["Дата", "Время", "ID промоутера", "Имя промоутера", "Адрес работы", "Планируемое время работы",
                 "Статус", "ID записи"])
            for r in reports:
                report_sheet.append([
                    r["Дата"], r["Время"], r["ID промоутера"], r["Имя промоутера"],
                    r["Адрес работы"], r["Планируемое время работы"], r["Статус"], r["ID записи"]
                ])

            report_file_name = f"report_{start_date}_{end_date}.xlsx"
            report_workbook.save(report_file_name)

            with open(report_file_name, 'rb') as f:
                await callback_query.message.answer_document(
                    types.BufferedInputFile(f.read(), filename=report_file_name),
                    caption=f"Отчет за период с {start_date} по {end_date}",
                    reply_markup=get_manager_keyboard())

            os.remove(report_file_name)

    elif command_type == "payments":
        reports = get_reports_for_period(start_date, end_date)
        payments = calculate_payments(reports)

        if not payments:
            await callback_query.message.answer(
                f"За выбранный период ({start_date} - {end_date}) подтвержденных выходов не найдено для расчета выплат.",
                reply_markup=get_manager_keyboard())
        else:
            payment_message = f"Расчет выплат за период с {start_date} по {end_date}:\n\n"
            for promoter, amount in payments.items():
                payment_message += f"- {promoter}: {amount} руб.\n"
            await callback_query.message.answer(payment_message, reply_markup=get_manager_keyboard())

    await state.clear()
    await callback_query.answer()


async def main():
    # Инициализация Excel файла при запуске бота
    init_excel()

    # Создаем объекты бота и диспетчера
    default_properties = DefaultBotProperties(parse_mode="HTML")
    bot = Bot(token=API_TOKEN, default=default_properties)
    dp = Dispatcher()

    # Регистрируем наш роутер (содержащий все обработчики) в основной диспетчер
    dp.include_router(router)

    logger.info("Starting bot...")
    await dp.start_polling(bot)


if __name__ == '__main__':
    # Запускаем бота
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Bot stopped by KeyboardInterrupt")
    except Exception as e:
        logger.error(f"Bot stopped with an error: {e}")